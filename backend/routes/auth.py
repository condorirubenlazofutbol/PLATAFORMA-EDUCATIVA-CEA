from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from pydantic import BaseModel, EmailStr
from typing import Optional
from database import get_db_connection
import security as auth, psycopg2, io, openpyxl

router = APIRouter()
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/auth/login")

class RegistroUsuario(BaseModel):
    nombre: str
    apellido: str
    carnet: str
    rol: str
    subsistema_id: int
    nivel_asignado: Optional[str] = None
    carrera_id: Optional[int] = None
    turno: Optional[str] = "Noche"

class TokenData(BaseModel):
    username: str | None = None

def rows_to_dicts(cursor, rows):
    cols = [d[0] for d in cursor.description]
    return [dict(zip(cols, r)) for r in rows]

@router.post("/login")
def login(form_data: OAuth2PasswordRequestForm = Depends()):
    try:
        conn = get_db_connection()
        if not conn:
            raise HTTPException(status_code=500, detail="Database connection error")
        cur = conn.cursor()
        cur.execute("SELECT id, nombre, email, password, rol, nivel_asignado, estado, subsistema_id FROM usuarios WHERE email=%s", (form_data.username,))
        row = cur.fetchone()
        cur.close(); conn.close()
        
        if not row:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
            
        estado = row[6]
        if estado == 'pausado':
            raise HTTPException(status_code=403, detail="Tu cuenta estÃ¡ suspendida por falta de pago o por el administrador.")
            
        is_valid = False
        try:
            is_valid = auth.verify_password(form_data.password, row[3])
        except Exception as ve:
            raise HTTPException(status_code=500, detail=f"Error verificando password: {str(ve)}")

        if not is_valid:
            raise HTTPException(status_code=401, detail="ContraseÃ±a incorrecta")

        token = auth.create_access_token(data={"sub": row[2]})
        rol_retornado = row[4]
        # Normalizar roles: unificar variantes para que el frontend funcione correctamente
        if rol_retornado == 'profesor':
            rol_retornado = 'docente'
        elif rol_retornado == 'jefe':
            rol_retornado = 'jefe_carrera'
        
        return {"access_token": token, "token_type": "bearer",
                "rol": rol_retornado, "nombre": row[1], "nivel_asignado": row[5], "subsistema_id": row[7]}
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Login crash: {str(e)}")


@router.get("/me")
def get_current_user(token: str = Depends(oauth2_scheme)):
    payload = auth.decode_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Token invÃ¡lido")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Database error")
    cur = conn.cursor()
    cur.execute("SELECT id, nombre, apellido, email, rol, nivel_asignado, subsistema_id FROM usuarios WHERE email=%s", (payload.get("sub"),))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return {"id": row[0], "nombre": row[1], "apellido": row[2], "email": row[3], "rol": row[4], "nivel_asignado": row[5], "subsistema_id": row[6]}

def generate_cea_email(nombre: str, apellido: str):
    # Usar solo primer nombre y primer apellido (paterno)
    clean_n = nombre.strip().split(" ")[0].lower()
    clean_a = apellido.strip().split(" ")[0].lower()
    return f"{clean_n}{clean_a}" + "@ceapailon.com"

def obtener_paralelo_disponible(cur, carrera_id, nivel, area, turno):
    limite = 30 if str(area).lower() == 'tÃ©cnica' else 40
    cur.execute('''
        SELECT paralelo, COUNT(*) 
        FROM inscripciones 
        WHERE carrera_id = %s AND nivel = %s AND turno = %s
        GROUP BY paralelo
        ORDER BY paralelo ASC
    ''', (carrera_id, nivel, turno))
    paralelos = cur.fetchall()
    
    if not paralelos: return 'A'
    for paralelo, count in paralelos:
        if count < limite: return paralelo
    ultimo_paralelo = paralelos[-1][0]
    return chr(ord(ultimo_paralelo) + 1) if len(ultimo_paralelo) == 1 and ultimo_paralelo.isalpha() else 'A'

@router.post("/register-usuario", dependencies=[Depends(get_current_user)])
def register_usuario(data: RegistroUsuario):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Error de base de datos")
    try:
        cur = conn.cursor()
        email = generate_cea_email(data.nombre, data.apellido)
        hashed = auth.get_password_hash(str(data.carnet).strip())
        
        # Verificar si ya existe el carnet
        cur.execute("SELECT id FROM usuarios WHERE carnet=%s", (str(data.carnet).strip(),))
        if cur.fetchone():
            raise HTTPException(status_code=400, detail="El carnet ya estÃ¡ registrado.")

        cur.execute(
            "INSERT INTO usuarios (subsistema_id, nombre, apellido, email, password, rol, nivel_asignado, carnet) VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
            (data.subsistema_id or 1, data.nombre.strip(), data.apellido.strip(), email, hashed, data.rol, data.nivel_asignado, str(data.carnet).strip())
        )
        new_id = cur.fetchone()[0]
        
        # --- LÃ³gica de InscripciÃ³n Inteligente Pro ---
        if data.rol == "estudiante" and data.nivel_asignado:
            nivel_str = data.nivel_asignado
            
            # Ãrea TÃ©cnica (formato: "Carrera - Nivel")
            if " - " in nivel_str:
                parts = nivel_str.split(" - ")
                carrera_nombre = parts[0].strip()
                nivel_nombre = parts[1].strip()
                
                cur.execute("SELECT id, area FROM carreras WHERE nombre = %s AND area = 'TÃ©cnica'", (carrera_nombre,))
                c_row = cur.fetchone()
                if c_row:
                    c_id = c_row[0]
                    area = c_row[1]
                    paralelo = obtener_paralelo_disponible(cur, c_id, nivel_nombre, area, data.turno)
                    cur.execute("INSERT INTO inscripciones (usuario_id, carrera_id, nivel, paralelo, turno) VALUES (%s, %s, %s, %s, %s) ON CONFLICT (usuario_id, carrera_id, turno) DO UPDATE SET paralelo = EXCLUDED.paralelo", (new_id, c_id, nivel_nombre, paralelo, data.turno))
                    
                    # Matricular automÃ¡ticamente en sus 5 mÃ³dulos
                    cur.execute("SELECT id FROM modulos WHERE carrera_id = %s AND nivel = %s", (c_id, nivel_nombre))
                    modulos = cur.fetchall()
                    for mod in modulos:
                        cur.execute("INSERT INTO progreso (usuario_id, modulo_id, estado) VALUES (%s, %s, 'cursando') ON CONFLICT DO NOTHING", (new_id, mod[0]))
            
            # Ãrea HumanÃ­stica (formato: "Nivel")
            else:
                nivel_nombre = nivel_str.strip()
                cur.execute("SELECT id, area FROM carreras WHERE area = 'HumanÃ­stica'")
                carreras_hum = cur.fetchall()
                for c_row in carreras_hum:
                    c_id = c_row[0]
                    area = c_row[1]
                    paralelo = obtener_paralelo_disponible(cur, c_id, nivel_nombre, area, data.turno)
                    cur.execute("INSERT INTO inscripciones (usuario_id, carrera_id, nivel, paralelo, turno) VALUES (%s, %s, %s, %s, %s) ON CONFLICT (usuario_id, carrera_id, turno) DO UPDATE SET paralelo = EXCLUDED.paralelo", (new_id, c_id, nivel_nombre, paralelo, data.turno))
                    
                    # Matricular automÃ¡ticamente en sus 2 mÃ³dulos por materia
                    cur.execute("SELECT id FROM modulos WHERE carrera_id = %s AND nivel = %s", (c_id, nivel_nombre))
                    modulos = cur.fetchall()
                    for mod in modulos:
                        cur.execute("INSERT INTO progreso (usuario_id, modulo_id, estado) VALUES (%s, %s, 'cursando') ON CONFLICT DO NOTHING", (new_id, mod[0]))

        if data.carrera_id:
            cur.execute("SELECT area FROM carreras WHERE id = %s", (data.carrera_id,))
            c_area_row = cur.fetchone()
            area = c_area_row[0] if c_area_row else 'HumanÃ­stica'
            paralelo = obtener_paralelo_disponible(cur, data.carrera_id, data.nivel_asignado, area, data.turno)
            cur.execute(
                "INSERT INTO inscripciones (usuario_id, carrera_id, nivel, paralelo, turno) VALUES (%s, %s, %s, %s, %s) ON CONFLICT (usuario_id, carrera_id, turno) DO UPDATE SET paralelo = EXCLUDED.paralelo",
                (new_id, data.carrera_id, data.nivel_asignado, paralelo, data.turno)
            )
            
        conn.commit()
        return {"id": new_id, "email": email, "mensaje": f"Usuario {data.nombre} ({data.rol}) creado exitosamente"}
    except psycopg2.errors.UniqueViolation:
        conn.rollback()
        raise HTTPException(status_code=400, detail="El email generado ya existe. Contacte a soporte.")
    except HTTPException: raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close(); conn.close()

@router.get("/plantilla-estudiantes")
def descargar_plantilla_estudiantes():
    from fastapi.responses import StreamingResponse
    import io
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estudiantes"
    
    # Formato estandarizado: 3 columnas
    headers = ["Nombres", "Apellidos", "Carnet / CI"]
    ws.append(headers)
    
    # Datos de ejemplo
    ws.append(["Juan", "Perez", "12345678"])
    ws.append(["Maria", "Lopez", "87654321"])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Plantilla_Estudiantes_CEA.xlsx"}
    )

@router.get("/plantilla-docentes")
def descargar_plantilla_docentes():
    from fastapi.responses import StreamingResponse
    import io
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Docentes"
    
    # Formato estandarizado: 3 columnas
    headers = ["Nombres", "Apellidos", "Carnet / CI"]
    ws.append(headers)
    
    # Datos de ejemplo
    ws.append(["Carlos", "Mamani", "8877665"])
    ws.append(["Ana", "Suarez", "5544332"])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Plantilla_Docentes_CEA.xlsx"}
    )

@router.post("/importar-estudiantes-excel")
async def importar_estudiantes_excel(nivel: str, turno: str = "Noche", file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    if current_user["rol"] not in ["admin", "administrador", "director", "secretaria"]:
        raise HTTPException(403, "No autorizado")
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        registrados = 0
        errores = []
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Columnas: Nombre | Apellido | Carnet (3 columnas)
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Ignorar filas completamente vacÃ­as
            if not row or not row[0]: continue
            
            nombres   = str(row[0]).strip() if len(row) > 0 and row[0] else ""
            apellidos = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            carnet    = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            
            # Validar que al menos tenga nombre y carnet
            if not nombres or not carnet:
                errores.append(f"Fila ignorada: nombre o carnet vacÃ­o")
                continue
            
            email = generate_cea_email(nombres, apellidos)
            password = auth.get_password_hash(carnet)
            
            try:
                # Insertar usuario â€” ON CONFLICT actualiza en vez de fallar si ya existe
                cur.execute(
                    """INSERT INTO usuarios (subsistema_id, nombre, apellido, email, password, rol, nivel_asignado, carnet)
                       VALUES (1,%s,%s,%s,%s,'estudiante',%s,%s)
                       ON CONFLICT (email) DO UPDATE SET
                           nombre = EXCLUDED.nombre,
                           apellido = EXCLUDED.apellido,
                           nivel_asignado = EXCLUDED.nivel_asignado,
                           carnet = EXCLUDED.carnet
                       RETURNING id""",
                    (nombres, apellidos, email, password, nivel, carnet)
                )
                new_id = cur.fetchone()[0]
                
                # --- LÃ³gica de InscripciÃ³n Inteligente Pro ---
                if " - " in nivel:
                    # Ãrea TÃ©cnica
                    parts = nivel.split(" - ")
                    carrera_nombre = parts[0].strip()
                    nivel_nombre = parts[1].strip()
                    
                    cur.execute("SELECT id, area FROM carreras WHERE nombre = %s AND area = 'TÃ©cnica'", (carrera_nombre,))
                    c_row = cur.fetchone()
                    if c_row:
                        c_id = c_row[0]
                        db_area = c_row[1]
                        paralelo = obtener_paralelo_disponible(cur, c_id, nivel_nombre, db_area, turno)
                        cur.execute("INSERT INTO inscripciones (usuario_id, carrera_id, nivel, paralelo, turno) VALUES (%s, %s, %s, %s, %s) ON CONFLICT (usuario_id, carrera_id, turno) DO UPDATE SET paralelo = EXCLUDED.paralelo", (new_id, c_id, nivel_nombre, paralelo, turno))
                        
                        # Matricular en mÃ³dulos
                        cur.execute("SELECT id FROM modulos WHERE carrera_id = %s AND nivel = %s", (c_id, nivel_nombre))
                        modulos = cur.fetchall()
                        for mod in modulos:
                            cur.execute("INSERT INTO progreso (usuario_id, modulo_id, estado) VALUES (%s, %s, 'cursando') ON CONFLICT DO NOTHING", (new_id, mod[0]))
                else:
                    # Ãrea HumanÃ­stica
                    nivel_nombre = nivel.strip()
                    cur.execute("SELECT id, area FROM carreras WHERE area = 'HumanÃ­stica'")
                    carreras_hum = cur.fetchall()
                    for c_row in carreras_hum:
                        c_id = c_row[0]
                        db_area = c_row[1]
                        paralelo = obtener_paralelo_disponible(cur, c_id, nivel_nombre, db_area, turno)
                        cur.execute("INSERT INTO inscripciones (usuario_id, carrera_id, nivel, paralelo, turno) VALUES (%s, %s, %s, %s, %s) ON CONFLICT (usuario_id, carrera_id, turno) DO UPDATE SET paralelo = EXCLUDED.paralelo", (new_id, c_id, nivel_nombre, paralelo, turno))
                        
                        # Matricular en mÃ³dulos
                        cur.execute("SELECT id FROM modulos WHERE carrera_id = %s AND nivel = %s", (c_id, nivel_nombre))
                        modulos = cur.fetchall()
                        for mod in modulos:
                            cur.execute("INSERT INTO progreso (usuario_id, modulo_id, estado) VALUES (%s, %s, 'cursando') ON CONFLICT DO NOTHING", (new_id, mod[0]))
                
                conn.commit() # Commit PER ROW to prevent losing previous rows on error
                registrados += 1
            except Exception as e:
                conn.rollback() # Solo deshace esta iteraciÃ³n porque las anteriores ya se hicieron commit
                errores.append(f"Error en CI {carnet}: {str(e)}")
                continue
            
        cur.close(); conn.close()
        
        return {"registrados": registrados, "errores": errores}
    except Exception as e:
        raise HTTPException(500, f"Error procesando Excel: {str(e)}")


@router.get("/estudiantes", dependencies=[Depends(get_current_user)])
def get_estudiantes():
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Error de base de datos")
    try:
        cur = conn.cursor()
        cur.execute("SELECT id, nombre, apellido, email, nivel_asignado, carnet, estado FROM usuarios WHERE rol='estudiante' ORDER BY nivel_asignado, nombre")
        return {"estudiantes": rows_to_dicts(cur, cur.fetchall())}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

@router.get("/profesores")
def get_profesores():
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Error de base de datos")
    try:
        cur = conn.cursor()
        cur.execute("SELECT id, nombre, apellido, email, nivel_asignado, carnet, estado FROM usuarios WHERE rol='profesor' ORDER BY nivel_asignado")
        return {"profesores": rows_to_dicts(cur, cur.fetchall())}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.get("/usuarios")
def get_usuarios(current_user: dict = Depends(get_current_user)):
    """Lista todos los usuarios para la SecretarÃ­a/Director/Admin."""
    if current_user["rol"] not in ["admin", "administrador", "director", "secretaria"]:
        raise HTTPException(status_code=403, detail="No autorizado")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Error de base de datos")
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT u.id, u.nombre, u.apellido, u.email, u.rol, u.nivel_asignado, 
                   u.carnet, u.estado, u.fecha_registro
            FROM usuarios u
            ORDER BY u.rol, u.nombre
        """)
        return rows_to_dicts(cur, cur.fetchall())
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.get("/inscripciones")
def get_inscripciones(current_user: dict = Depends(get_current_user)):
    """Lista todas las inscripciones con datos del estudiante, carrera y paralelo."""
    if current_user["rol"] not in ["admin", "administrador", "director", "secretaria", "jefe_carrera"]:
        raise HTTPException(status_code=403, detail="No autorizado")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Error de base de datos")
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT 
                i.id, i.usuario_id, i.carrera_id, i.nivel, i.paralelo, i.estado, i.fecha_inscripcion,
                u.nombre, u.apellido, u.carnet, u.email,
                c.nombre AS carrera_nombre, c.area AS carrera_area
            FROM inscripciones i
            JOIN usuarios u ON u.id = i.usuario_id
            JOIN carreras c ON c.id = i.carrera_id
            ORDER BY c.area, c.nombre, i.nivel, i.paralelo, u.apellido
        """)
        return {"inscripciones": rows_to_dicts(cur, cur.fetchall())}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.put("/inscripciones/{inscripcion_id}/estado")
def cambiar_estado_inscripcion(inscripcion_id: int, current_user: dict = Depends(get_current_user)):
    """Cambia el estado de una inscripciÃ³n (activo/pausado/retirado)."""
    if current_user["rol"] not in ["admin", "administrador", "director", "secretaria"]:
        raise HTTPException(status_code=403, detail="No autorizado")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Error de base de datos")
    try:
        from pydantic import BaseModel as BM
        # State toggle: activo -> pausado -> retirado -> activo
        cur = conn.cursor()
        cur.execute("SELECT estado FROM inscripciones WHERE id=%s", (inscripcion_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "InscripciÃ³n no encontrada")
        next_state = {"activo": "pausado", "pausado": "retirado", "retirado": "activo"}.get(row[0], "activo")
        cur.execute("UPDATE inscripciones SET estado=%s WHERE id=%s", (next_state, inscripcion_id))
        conn.commit()
        return {"nuevo_estado": next_state}
    finally:
        conn.close()

@router.get("/personal")
def get_personal(current_user: dict = Depends(get_current_user)):
    if current_user["rol"] not in ["admin", "administrador", "director", "secretaria"]:
        raise HTTPException(status_code=403, detail="No autorizado")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Error DB")
    try:
        cur = conn.cursor()
        if current_user["rol"] in ("admin", "administrador"):
            # Admin ve TODOS excepto otros admins
            cur.execute("""
                SELECT id, nombre, apellido, email, rol, nivel_asignado, carnet, estado,
                       COALESCE(es_jefe, FALSE) as es_jefe
                FROM usuarios
                WHERE rol NOT IN ('admin', 'administrador')
                ORDER BY
                    CASE rol
                        WHEN 'director' THEN 1
                        WHEN 'secretaria' THEN 2
                        WHEN 'jefe_carrera' THEN 3
                        WHEN 'docente' THEN 4
                        WHEN 'profesor' THEN 4
                        ELSE 5
                    END, apellido, nombre
            """)
        else:
            cur.execute("""
                SELECT id, nombre, apellido, email, rol, nivel_asignado, carnet, estado,
                       COALESCE(es_jefe, FALSE) as es_jefe
                FROM usuarios
                WHERE rol IN ('profesor','docente','jefe_carrera','secretaria','director')
                ORDER BY rol, nombre
            """)
        return {"personal": rows_to_dicts(cur, cur.fetchall())}
    finally:
        conn.close()

@router.delete("/usuarios/{usuario_id}")
def delete_usuario(usuario_id: int, current_user: dict = Depends(get_current_user)):
    """Elimina un usuario del sistema.
    - Director: puede eliminar docentes y estudiantes.
    - Admin: puede eliminar cualquier usuario (excepto otros admins).
    - No se permite eliminar Director o Admin por parte de roles inferiores.
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Error de base de datos")
    try:
        cur = conn.cursor()
        # Verificar que el usuario existe y obtener su rol
        cur.execute("SELECT email, rol FROM usuarios WHERE id = %s", (usuario_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")

        rol_objetivo = row[1]
        mi_rol = current_user["rol"]

        # Protecciones de seguridad
        if rol_objetivo in ("admin", "administrador"):
            raise HTTPException(status_code=403, detail="No se puede eliminar una cuenta de Administrador.")
        if rol_objetivo == "director" and mi_rol not in ("admin", "administrador"):
            raise HTTPException(status_code=403, detail="Solo el Super Admin puede eliminar una cuenta de Director.")
        if mi_rol not in ("admin", "administrador", "director", "secretaria"):
            raise HTTPException(status_code=403, detail="No tiene permisos para eliminar usuarios.")

        cur.execute("DELETE FROM usuarios WHERE id = %s", (usuario_id,))
        conn.commit()
        return {"mensaje": "Usuario eliminado correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

class PasswordResetBody(BaseModel):
    new_password: str

@router.put("/usuarios/{usuario_id}/password", dependencies=[Depends(get_current_user)])
def reset_password(usuario_id: int, data: PasswordResetBody):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Error de base de datos")
    try:
        cur = conn.cursor()
        hashed = auth.get_password_hash(data.new_password)
        cur.execute("UPDATE usuarios SET password = %s WHERE id = %s", (hashed, usuario_id))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        conn.commit()
        return {"mensaje": "ContraseÃ±a restablecida correctamente"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

class EstadoUpdateBody(BaseModel):
    estado: str

@router.put("/usuarios/{usuario_id}/estado", dependencies=[Depends(get_current_user)])
def update_estado(usuario_id: int, data: EstadoUpdateBody):
    if data.estado not in ['activo', 'pausado']:
        raise HTTPException(status_code=400, detail="Estado invÃ¡lido")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Error de base de datos")
    try:
        cur = conn.cursor()
        cur.execute("UPDATE usuarios SET estado = %s WHERE id = %s", (data.estado, usuario_id))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        conn.commit()
        return {"mensaje": f"Estado actualizado a {data.estado}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

class EspecialidadUpdateBody(BaseModel):
    especialidad: str

@router.put("/usuarios/{usuario_id}/especialidad")
def update_especialidad(usuario_id: int, data: EspecialidadUpdateBody, current_user: dict = Depends(get_current_user)):
    """Asigna un curso/nivel a un docente con validación de unicidad:
    - Área Técnica: Solo un docente por nivel dentro de la misma especialidad (carrera).
    - Área Humanística: Solo un docente por materia/nivel.
    """
    if current_user["rol"] not in ["director", "admin", "administrador"]:
        raise HTTPException(status_code=403, detail="Solo el Director puede designar niveles a docentes.")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Error de base de datos")
    try:
        cur = conn.cursor()

        # 1. Obtener datos del docente que se quiere asignar
        cur.execute(
            "SELECT nivel_asignado, curso_asignado FROM usuarios WHERE id = %s AND rol IN ('docente', 'profesor', 'jefe_carrera')",
            (usuario_id,)
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Docente no encontrado")

        especialidad_docente = row[0] or ""  # Ej: "Sistemas Informáticos" o "Aplicados"
        nuevo_nivel = data.especialidad.strip()

        # 2. Determinar el área del docente buscando su especialidad en el catálogo de carreras
        cur.execute("SELECT area FROM carreras WHERE nombre = %s AND estado = 'activo' LIMIT 1", (especialidad_docente,))
        carrera_row = cur.fetchone()
        area_docente = (carrera_row[0] or "").lower() if carrera_row else "humanistica"
        # Fallback: si no está en carreras, es humanística (su especialidad es la materia)
        is_tecnica = "cnica" in area_docente  # cubre "Técnica" con encoding

        # 3. Validar conflicto de asignación
        # nuevo_nivel puede ser multi-materia: "Complementarios (Segundo Año), Especializados (Tercer Año)"
        # Descomponer en lista de materias/niveles individuales
        nuevos_niveles = [n.strip() for n in nuevo_nivel.split(",") if n.strip()]

        if is_tecnica:
            # Técnica: 1 docente por nivel dentro de la misma carrera
            # Obtener todos los demás docentes de esa carrera
            cur.execute("""
                SELECT u.id, u.nombre, u.apellido, u.curso_asignado
                FROM usuarios u
                WHERE u.id != %s
                  AND u.rol IN ('docente', 'profesor', 'jefe_carrera')
                  AND u.nivel_asignado = %s
                  AND u.estado = 'activo'
                  AND u.curso_asignado IS NOT NULL
            """, (usuario_id, especialidad_docente))
        else:
            # Humanística: 1 docente por materia DENTRO de la misma especialidad
            cur.execute("""
                SELECT u.id, u.nombre, u.apellido, u.curso_asignado
                FROM usuarios u
                WHERE u.id != %s
                  AND u.rol IN ('docente', 'profesor', 'jefe_carrera')
                  AND u.nivel_asignado = %s
                  AND u.estado = 'activo'
                  AND u.curso_asignado IS NOT NULL
            """, (usuario_id, especialidad_docente))

        otros_docentes = cur.fetchall()

        # Verificar intersección materia por materia
        for otro in otros_docentes:
            otro_id, otro_nom, otro_ap, otro_cursos = otro
            if not otro_cursos:
                continue
            # Niveles ya asignados al otro docente (puede ser multi)
            niveles_otro = [n.strip() for n in otro_cursos.split(",") if n.strip()]
            # Encontrar solapamiento
            solapamiento = [n for n in nuevos_niveles if n in niveles_otro]
            if solapamiento:
                nombre_conflicto = f"{otro_nom} {otro_ap}"
                materias = ", ".join(solapamiento)
                if is_tecnica:
                    detalle = (f"El nivel '{materias}' en {especialidad_docente} ya está asignado "
                               f"al docente {nombre_conflicto}. Libere ese nivel antes de asignarlo aquí.")
                else:
                    detalle = (f"La materia '{materias}' ya está asignada al docente {nombre_conflicto} "
                               f"en {especialidad_docente}. Cada materia solo puede tener un docente por especialidad.")
                raise HTTPException(status_code=409, detail=detalle)

        # 4. Sin conflicto: guardar la asignación
        cur.execute(
            "UPDATE usuarios SET curso_asignado = %s WHERE id = %s AND rol IN ('docente', 'profesor', 'jefe_carrera')",
            (nuevo_nivel, usuario_id)
        )
        conn.commit()
        return {"mensaje": f"Nivel/Curso asignado: {nuevo_nivel}"}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.put("/update-password")
def update_my_password(data: PasswordResetBody, current_user: dict = Depends(get_current_user)):
    """Permite al usuario logueado cambiar su propia contraseÃ±a."""
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        hashed = auth.get_password_hash(data.new_password)
        cur.execute("UPDATE usuarios SET password = %s WHERE id = %s", (hashed, current_user["id"]))
        conn.commit()
        return {"mensaje": "Tu contraseÃ±a ha sido actualizada correctamente"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.post("/bulk-register", dependencies=[Depends(get_current_user)])
async def bulk_register(nivel: str, turno: str = "Noche", rol: str = "estudiante", file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Carga masiva de usuarios (estudiante o profesor) desde Excel (.xlsx)."""
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx")
    
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    sheet = wb.active
    
    registrados = 0
    errores = []
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    subsistema_id = current_user.get("subsistema_id") or 1
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        nombre = str(row[0]).strip() if row[0] else ""
        carnet = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        if not nombre or not carnet: continue
        
        apellido = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        
        try:
            clean_n = nombre.split(' ')[0].lower()
            clean_a = apellido.split(' ')[0].lower()
            email = f"{clean_n}{clean_a}@ceapailon.com"
            
            s_carnet = str(carnet).strip()
            hashed = auth.get_password_hash(s_carnet)
            
            # Rol para docente: usar 'docente'
            db_rol = 'docente' if rol in ('docente', 'profesor') else rol
            
            cur.execute(
                """INSERT INTO usuarios (subsistema_id, nombre, apellido, email, password, rol, nivel_asignado, carnet)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                   ON CONFLICT (email) DO UPDATE SET
                       nombre = EXCLUDED.nombre,
                       apellido = EXCLUDED.apellido,
                       nivel_asignado = EXCLUDED.nivel_asignado,
                       carnet = EXCLUDED.carnet
                   RETURNING id""",
                (subsistema_id, nombre, apellido, email, hashed, db_rol, nivel, s_carnet)
            )
            new_id = cur.fetchone()[0]

            # --- Lógica de Inscripción Inteligente Pro (Carga Masiva) ---
            if db_rol == "estudiante" and nivel:
                nivel_str = nivel
                # Área Técnica
                if " - " in nivel_str:
                    parts = nivel_str.split(" - ")
                    carrera_nombre = parts[0].strip()
                    nivel_nombre = parts[1].strip()
                    cur.execute("SELECT id, area FROM carreras WHERE nombre = %s AND area = 'TÃ©cnica'", (carrera_nombre,))
                    c_row = cur.fetchone()
                    if c_row:
                        c_id, area = c_row
                        paralelo = obtener_paralelo_disponible(cur, c_id, nivel_nombre, area, turno)
                        cur.execute("INSERT INTO inscripciones (usuario_id, carrera_id, nivel, paralelo, turno) VALUES (%s, %s, %s, %s, %s) ON CONFLICT (usuario_id, carrera_id, turno) DO UPDATE SET paralelo = EXCLUDED.paralelo", (new_id, c_id, nivel_nombre, paralelo, turno))
                # Ã rea HumanÃ­stica
                else:
                    nivel_nombre = nivel_str.strip()
                    cur.execute("SELECT id, area FROM carreras WHERE area = 'HumanÃ­stica'")
                    for c_id, area in cur.fetchall():
                        paralelo = obtener_paralelo_disponible(cur, c_id, nivel_nombre, area, turno)
                        cur.execute("INSERT INTO inscripciones (usuario_id, carrera_id, nivel, paralelo, turno) VALUES (%s, %s, %s, %s, %s) ON CONFLICT (usuario_id, carrera_id, turno) DO UPDATE SET paralelo = EXCLUDED.paralelo", (new_id, c_id, nivel_nombre, paralelo, turno))

            conn.commit()
            registrados += 1
        except Exception as e:
            conn.rollback()
            errores.append(f"Error en fila {nombre}: {str(e)}")
            continue

    cur.close(); conn.close()
    
    return {
        "status": "done",
        "registrados": registrados,
        "errores": errores
    }

class PromoverDirectorRequest(BaseModel):
    usuario_id: int
    nuevo_rol: str  # 'director', 'secretaria', 'jefe_carrera', 'docente'

@router.post("/promover-alta-direccion")
def promover_alta_direccion(data: PromoverDirectorRequest, current_user: dict = Depends(get_current_user)):
    """Admin puede asignar director, secretaria, jefe_carrera, docente.
    - Solo 1 director y 1 secretaria en todo el sistema.
    - Si ya existe uno, se degrada a 'docente' automáticamente.
    """
    if current_user["rol"] not in ["admin", "administrador"]:
        raise HTTPException(status_code=403, detail="Solo el Súper Admin puede realizar esta acción")
    if data.nuevo_rol not in ["director", "secretaria", "jefe_carrera", "docente", "profesor"]:
        raise HTTPException(status_code=400, detail="Rol inválido")

    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Error DB")
    try:
        cur = conn.cursor()

        # Verificar que el usuario objetivo existe y no es admin
        cur.execute("SELECT nombre, apellido, rol FROM usuarios WHERE id = %s", (data.usuario_id,))
        target = cur.fetchone()
        if not target:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        if target[2] in ("admin", "administrador"):
            raise HTTPException(status_code=403, detail="No se puede modificar el rol de un Administrador")

        anterior_nombre = None

        # Para director y secretaria: solo puede haber 1
        if data.nuevo_rol in ("director", "secretaria"):
            cur.execute(
                "SELECT id, nombre, apellido FROM usuarios WHERE rol = %s AND id != %s",
                (data.nuevo_rol, data.usuario_id)
            )
            anterior = cur.fetchone()
            if anterior:
                anterior_nombre = f"{anterior[1]} {anterior[2]}"
                # Degradar al anterior a docente
                cur.execute("UPDATE usuarios SET rol = 'docente' WHERE id = %s", (anterior[0],))

        # Si el usuario era jefe y cambia a otro rol, retirar es_jefe
        cur.execute("UPDATE usuarios SET rol = %s, es_jefe = FALSE WHERE id = %s",
                    (data.nuevo_rol, data.usuario_id))

        conn.commit()

        msg = f"{target[0]} {target[1]} ahora es {data.nuevo_rol}."
        if anterior_nombre:
            msg += f" {anterior_nombre} fue degradado/a a docente."
        return {"mensaje": msg}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

class PromoverJefeRequest(BaseModel):
    usuario_id: int
    carrera_id: Optional[int] = None
    especialidad_nombre: Optional[str] = None  # Nombre de la especialidad/materia del grupo

@router.post("/promover-jefe-carrera")
def promover_jefe_carrera(data: PromoverJefeRequest, current_user: dict = Depends(get_current_user)):
    """Designa a un docente como Jefe de su especialidad/carrera.
    
    Reglas:
    - Solo 1 Jefe por especialidad (Humanística) o por carrera (Técnica).
    - Si ya hay un Jefe en esa especialidad, se le retira la flag es_jefe automáticamente.
    - El ROL del docente permanece como 'docente' — solo cambia es_jefe=True.
    - Solo el Director puede designar Jefes.
    """
    if current_user["rol"] not in ["admin", "administrador", "director"]:
        raise HTTPException(status_code=403, detail="Solo la Dirección puede nombrar Jefes de Carrera")

    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Error DB")
    try:
        cur = conn.cursor()

        # 1. Obtener la especialidad del docente a promover (nivel_asignado = su área/materia)
        cur.execute(
            "SELECT nombre, apellido, nivel_asignado, rol FROM usuarios WHERE id = %s",
            (data.usuario_id,)
        )
        docente = cur.fetchone()
        if not docente:
            raise HTTPException(status_code=404, detail="Docente no encontrado")

        nombre_docente = f"{docente[0]} {docente[1]}"
        especialidad_docente = docente[2] or data.especialidad_nombre or ""

        if not especialidad_docente:
            raise HTTPException(status_code=400, detail="El docente no tiene una especialidad asignada.")

        # 2. Retirar la flag es_jefe a cualquier otro docente de LA MISMA ESPECIALIDAD
        #    (garantiza que solo 1 sea Jefe por especialidad)
        cur.execute("""
            UPDATE usuarios
            SET es_jefe = FALSE
            WHERE id != %s
              AND nivel_asignado = %s
              AND es_jefe = TRUE
              AND rol IN ('docente', 'profesor', 'jefe_carrera')
        """, (data.usuario_id, especialidad_docente))

        # 3. Actualizar carreras.jefe_id si se envió carrera_id o se puede resolver
        carrera_id = data.carrera_id
        if not carrera_id and data.especialidad_nombre:
            cur.execute("SELECT id FROM carreras WHERE nombre ILIKE %s LIMIT 1", (data.especialidad_nombre.strip(),))
            c_row = cur.fetchone()
            if c_row:
                carrera_id = c_row[0]

        if carrera_id:
            cur.execute("UPDATE carreras SET jefe_id = %s WHERE id = %s", (data.usuario_id, carrera_id))

        # 4. Marcar al docente como Jefe con es_jefe=True (SIN cambiar su rol)
        cur.execute("""
            UPDATE usuarios
            SET es_jefe = TRUE
            WHERE id = %s
        """, (data.usuario_id,))

        conn.commit()
        return {"mensaje": f"{nombre_docente} ha sido designado Jefe de {especialidad_docente} exitosamente."}
    except HTTPException:
        raise
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")
    finally:
        if conn:
            conn.close()

@router.post("/retirar-jefe-carrera")
def retirar_jefe_carrera(data: PromoverJefeRequest, current_user: dict = Depends(get_current_user)):
    """Retira la designación de Jefe a un docente sin cambiar su rol de docente."""
    if current_user["rol"] not in ["admin", "administrador", "director"]:
        raise HTTPException(status_code=403, detail="Solo la Dirección puede gestionar Jefes de Carrera")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Error DB")
    try:
        cur = conn.cursor()
        cur.execute("UPDATE usuarios SET es_jefe = FALSE WHERE id = %s", (data.usuario_id,))
        if data.carrera_id:
            # Limpiar jefe_id en carreras si corresponde
            cur.execute("UPDATE carreras SET jefe_id = NULL WHERE id = %s AND jefe_id = %s",
                       (data.carrera_id, data.usuario_id))
        conn.commit()
        return {"mensaje": "Designación de Jefe retirada correctamente."}
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn:
            conn.close()

